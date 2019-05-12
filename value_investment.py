from openpyxl import load_workbook
import os
import re
from openpyxl.styles import Border, Side, PatternFill
from datetime import datetime


class ValuePrice(object):
    def __init__(self):
        self.pe_path = 'data/pe'
        self.report_path = 'data/report'
        self.stock_info = []
        self.key = None
        self.template = 'template/template_value_investment.xlsx'

    @staticmethod
    def get_code(file):
        codes = []
        f = open(file, 'r')
        lines = f.readlines()
        f.close()
        for line in lines:
            codes.append(line.strip())
        return codes

    def iterator_code(self, file):
        codes = self.get_code(file)
        for code in codes:
            yield code

    @staticmethod
    def format_result(num):
        num_x, num_y = str(num).split('.')
        num = float(num_x + '.' + num_y[0:2])
        return num

    def estimate_value_price(self, code):
        stock_info = None
        file = os.path.join(self.pe_path, code + '.xlsx')
        report_excel = os.path.join(self.report_path, code + '.xlsx')
        roes = []
        wb = load_workbook(report_excel)
        ws = wb['report_year']
        for i in range(5):
            row_roe = chr(98 + i) + '21'
            roes.append(ws[row_roe].value)
        wb.close()
        while True:
            if os.path.exists(file):
                wb = load_workbook(file)
                print(code)
                ws_pe = wb['pe']
                industy = ws_pe['d1'].value
                row_num = ws_pe.max_row
                pe = ws_pe['b2'].value
                pe = float(str(pe).replace(',', ''))
                if pe == 0:
                    break
                price = ws_pe['f1'].value
                price_num = re.findall(r'\d+', str(price))
                if len(price_num) == 2:
                    price = price_num[0] + '.' + price_num[1]
                else:
                    price = price_num[0]
                price = float(price)
                eps = self.format_result(price / pe)
                stock_name = ws_pe['b1'].value
                stock_name = str(stock_name).split('(')[0]
                pe_low_avg_row = 'd' + str(row_num)
                pe_low_avg = ws_pe[pe_low_avg_row].value
                pe_low_avg = float(pe_low_avg)
                if pe_low_avg == 0:
                    break
                pe_avg_row = 'b' + str(row_num)
                pe_avg = ws_pe[pe_avg_row].value
                pe_avg = float(pe_avg)
                pe_high_avg_row = 'c' + str(row_num)
                pe_high_avg = ws_pe[pe_high_avg_row].value
                pe_high_avg = float(pe_high_avg)
                low_value_price = eps * pe_low_avg
                low_value_price = self.format_result(low_value_price)
                # low_value_price = round(low_value_price, 2)
                avg_value_price = eps * pe_avg
                avg_value_price = self.format_result(avg_value_price)
                # avg_value_price = round(avg_value_price, 2)
                high_value_price = eps * pe_high_avg
                high_value_price = self.format_result(high_value_price)
                # high_value_price = round(high_value_price, 2)
                now_to_low_rate = (pe_low_avg - pe) / pe * 100
                now_to_low_rate = round(now_to_low_rate)
                low_to_normal_rate = (pe_avg - pe_low_avg) / pe_low_avg * 100
                low_to_normal_rate = round(low_to_normal_rate, 2)
                dividend_rate = ws_pe['f2'].value
                dividend_rate = float(str(dividend_rate).replace('%', ''))
                pb = ws_pe['d2'].value
                pb = float(pb)
                # market_value = ws_pe['j2'].value
                market_value = ws_pe['h2'].value
                market_value = float(re.sub(r'[,亿]', '', str(market_value)).strip())
                ws_fcf = wb['cash_flow']
                now_fcf = ws_fcf['b3'].value
                now_fcf = round(float(now_fcf), 2)
                stock_info = [code, stock_name, industy, eps, price, low_value_price, avg_value_price, high_value_price,
                              pe, pe_low_avg, pe_avg, pe_high_avg, now_to_low_rate, low_to_normal_rate, dividend_rate,
                              pb, market_value, now_fcf]
                ws_fcf_row_max = ws_fcf.max_row
                value_count = 0
                for i in range(3, ws_fcf_row_max + 1):
                    row_date = 'a' + str(i)
                    date_year = ws_fcf[row_date].value
                    if '12-31' in date_year:
                        row_value = 'b' + str(i)
                        fcf = float(ws_fcf[row_value].value)
                        roe = roes[0]
                        roes.remove(roe)
                        roe = float(roe)
                        if roe > 20 and fcf > 0:
                            stock_info.append('A级')
                        elif roe > 20 and fcf < 0:
                            stock_info.append('B-级')
                        elif 10 < roe < 20 and fcf > 0:
                            stock_info.append('B+级')
                        elif 10 < roe < 20 and fcf < 0:
                            stock_info.append('B-级')
                        elif 0 < roe < 10 and fcf > 0:
                            stock_info.append('C级')
                        elif 10 > roe > 0 > fcf:
                            stock_info.append('D级')
                        else:
                            stock_info.append('垃圾级')
                        value_count += 1
                        if value_count > 4:
                            break
                # print(stock_info)
                self.stock_info.append(stock_info)
                break
            else:
                break
        return stock_info

    def write_excle(self, wb, key):
        sheets = {'A级': 'A', 'B+级': 'B+', 'B-级': 'B-', 'C级': 'C', 'D级': 'D', }
        ws = wb[sheets[key]]
        for i in range(len(self.stock_info)):
            for j in range(len(self.stock_info[i])):
                row = chr(97 + j) + str(i + 2)
                ws[row] = self.stock_info[i][j]
        bd = Side(style='thin', color='000000')
        for m in range(97, 97 + len(self.stock_info[0])):
            for n in range(0, len(self.stock_info) + 1):
                row = chr(m) + str(n + 1)
                p = ws[row]
                p.border = Border(top=bd, left=bd, right=bd, bottom=bd)

    def main(self, file, wb):
        self.key = re.split(r'[/.\\]', str(file))[2]
        for code in self.iterator_code(file):
            self.estimate_value_price(code)
        self.write_excle(wb, self.key)
        self.stock_info = []

    @staticmethod
    def mark():
        wb = load_workbook('value_investment_%s.xlsx' % date)
        fill = PatternFill("solid", fgColor="7CCD7C")
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            row_num = ws.max_row
            for i in range(2, row_num):
                pe_row = 'i' + str(i)
                pe = ws[pe_row].value
                low_pe_row = 'j' + str(i)
                low_pe = ws[low_pe_row].value
                if pe < low_pe:
                    name_row = 'b' + str(i)
                    row = ws[name_row]
                    row.fill = fill
        wb.save('value_investment_%s.xlsx' % date)


if __name__ == '__main__':
    date = datetime.now().strftime('%Y%m%d')
    app = ValuePrice()
    # app.estimate_value_price('000001')
    workbook = load_workbook(app.template)
    for stock_file in os.listdir('data/stock_classify'):
        path = os.path.join('data/stock_classify', stock_file)
        app.main(path, workbook)
    workbook.save('value_investment_%s.xlsx' % date)
    app.mark()
