from fun.url_parser import UrlParser
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from multiprocessing import Pool
import json


class ENiu(object):
    def __init__(self):
        self.urlparser = UrlParser()
        self.pe_folder = './data/pe'
        self.template = 'template/template_eniu.xlsx'
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
        }

    @staticmethod
    def get_stock_codes():
        codes = []
        with open('stock.txt', 'r') as f:
            lines = f.readlines()
        for line in lines:
            codes.append(str(line).strip())
        return codes

    def iterator_stock_code(self):
        codes = self.get_stock_codes()
        for code in codes:
            yield code

    def get_pe_info(self, code):
        stock_fcf = self.get_free_cash_flow(code)
        stock_info_pe = {}
        line_number = None
        if '6' in code[0]:
            code = 'sh%s' % code
        else:
            code = 'sz%s' % code
        stock_url = 'https://eniu.com/gu/%s' % code
        soup = self.urlparser.soup_request(stock_url, headers=self.headers)
        # print(soup)
        html = soup.find_all('div', class_='col-xs-12')
        # print(html)
        # print(len(html))
        if len(html) == 0:
            print("No page")
        else:
            if len(html) == 11:
                # html_pe = html[2]
                html_pe = html[6]
            else:
                # html_pe = html[3]
                html_pe = html[7]
            # print(html_pe)
            html_fundamentals = html[0]
            soup_fundamentals = self.urlparser.lxml_html(html_fundamentals)
            fundamentals_info = soup_fundamentals.find_all('a')
            stock_name = fundamentals_info[0].string
            stock_info_pe['b1'] = stock_name
            stock_price = fundamentals_info[2].string
            stock_info_pe['f1'] = stock_price
            stock_pe = fundamentals_info[3].string
            stock_info_pe['b2'] = stock_pe
            stock_pb = fundamentals_info[4].string
            stock_info_pe['d2'] = stock_pb
            stock_dy = fundamentals_info[5].string
            stock_info_pe['f2'] = stock_dy
            stock_roe = fundamentals_info[6].string
            stock_info_pe['h2'] = stock_roe
            stock_mv = fundamentals_info[7].string
            stock_info_pe['j2'] = stock_mv
            soup_pe = self.urlparser.lxml_html(html_pe)
            # print(soup_pe)
            pe_info = soup_pe.find_all('h3')
            # print(pe_info)
            # pe_now = pe_info[2].string
            pe_now = stock_pe
            stock_info_pe['b4'] = pe_now
            pe_avg = pe_info[2].string
            stock_info_pe['d4'] = pe_avg
            pe_high = pe_info[3].string
            stock_info_pe['f4'] = pe_high
            pe_low = pe_info[4].string
            stock_info_pe['h4'] = pe_low
            pe_his_avg = []
            pe_his_high = []
            pe_his_low = []
            html_pe_history = soup_pe.find_all('td')

            def get_pe_history(num_1):
                date = html_pe_history[0 + 4 * num_1].string
                pe_history_avg = html_pe_history[1 + 4 * num_1].string
                pe_history_high = html_pe_history[2 + 4 * num_1].string
                pe_history_low = html_pe_history[3 + 4 * num_1].string
                if pe_history_avg is None:
                    pe_history_avg = 0
                if pe_history_high is None:
                    pe_history_high = 0
                if pe_history_low is None:
                    pe_history_low = 0
                pe_history_avg = str(pe_history_avg).replace(',', '')
                pe_history_high = str(pe_history_high).replace(',', '')
                pe_history_low = str(pe_history_low).replace(',', '')
                pe_his_avg.append(pe_history_avg)
                pe_his_high.append(pe_history_high)
                pe_his_low.append(pe_history_low)
                col_date = 'a' + str(7 + num_1)
                stock_info_pe[col_date] = date
                col_pe_history_avg = 'b' + str(7 + num_1)
                stock_info_pe[col_pe_history_avg] = float(pe_history_avg)
                col_pe_history_high = 'c' + str(7 + num_1)
                stock_info_pe[col_pe_history_high] = float(pe_history_high)
                col_pe_history_low = 'd' + str(7 + num_1)
                stock_info_pe[col_pe_history_low] = float(pe_history_low)

            # print(len(html_pe_history))
            # print(html_pe_history)
            if len(html_pe_history) > 40:
                html_pe_history = html_pe_history[:40]
                for i in range(0, 10):
                    get_pe_history(i)
                    stock_info_pe['b17'] = self.avg(pe_his_avg)
                    stock_info_pe['c17'] = self.avg(pe_his_high)
                    stock_info_pe['d17'] = self.avg(pe_his_low)
                line_number = 17
            else:
                number = int(len(html_pe_history) / 4)
                for i in range(0, int(len(html_pe_history) / 4)):
                    get_pe_history(i)
                row_avg = 'b' + str(17 - 10 + number)
                row_high = 'c' + str(17 - 10 + number)
                row_low = 'd' + str(17 - 10 + number)
                stock_info_pe[row_avg] = self.avg(pe_his_avg)
                stock_info_pe[row_high] = self.avg(pe_his_high)
                stock_info_pe[row_low] = self.avg(pe_his_low)
                line_number = 17 - 10 + number
            soup_corp = self.urlparser.lxml_html(html[-1])
            # print(soup_corp)
            corp_info = soup_corp.find_all('a')
            corp_industry = corp_info[0].string
            stock_info_pe['d1'] = corp_industry
            corp_page = corp_info[1].get('href')
            stock_info_pe['h1'] = corp_page
        stock_info_pb = self.get_pb_info(code, stock_info_pe)

        return stock_info_pe, stock_info_pb, stock_fcf, line_number

    def get_pb_info(self, code, stock_info_pe):
        stock_info_pb = stock_info_pe.copy()
        stock_url = 'https://eniu.com/gu/%s/pb' % code
        soup = self.urlparser.soup_request(stock_url)
        html = soup.find_all('div', class_='col-xs-12')
        # print(len(html))
        if len(html) == 0:
            print("No page")
        else:
            if len(html) == 10:
                # html_pb = html[2]
                html_pb = html[6]
            else:
                # html_pb = html[3]
                html_pb = html[7]
            soup_pb = self.urlparser.lxml_html(html_pb)
            pb_info = soup_pb.find_all('h3')
            pb_now = pb_info[1].string
            stock_info_pb['b4'] = pb_now
            pb_avg = pb_info[2].string
            stock_info_pb['d4'] = pb_avg
            pb_high = pb_info[3].string
            stock_info_pb['f4'] = pb_high
            pb_low = pb_info[4].string
            stock_info_pb['h4'] = pb_low
            pb_his_avg = []
            pb_his_high = []
            pb_his_low = []
            html_pb_history = soup_pb.find_all('td')

            def get_pb_history(num_1):
                date = html_pb_history[0 + 4 * num_1].string
                pb_history_avg = html_pb_history[1 + 4 * num_1].string
                pb_history_high = html_pb_history[2 + 4 * num_1].string
                pb_history_low = html_pb_history[3 + 4 * num_1].string
                if pb_history_avg is None:
                    pb_history_avg = 0
                if pb_history_high is None:
                    pb_history_high = 0
                if pb_history_low is None:
                    pb_history_low = 0
                pb_history_avg = str(pb_history_avg).replace(',', '')
                pb_history_high = str(pb_history_high).replace(',', '')
                pb_history_low = str(pb_history_low).replace(',', '')
                pb_his_avg.append(pb_history_avg)
                pb_his_high.append(pb_history_high)
                pb_his_low.append(pb_history_low)
                col_date = 'a' + str(7 + num_1)
                stock_info_pb[col_date] = date
                col_pb_history_avg = 'b' + str(7 + num_1)
                stock_info_pb[col_pb_history_avg] = float(pb_history_avg)
                col_pb_history_high = 'c' + str(7 + num_1)
                stock_info_pb[col_pb_history_high] = float(pb_history_high)
                col_pb_history_low = 'd' + str(7 + num_1)
                stock_info_pb[col_pb_history_low] = float(pb_history_low)

            if len(html_pb_history) > 40:
                html_pb_history = html_pb_history[:40]
                for i in range(0, 10):
                    get_pb_history(i)
                stock_info_pb['b17'] = self.avg(pb_his_avg)
                stock_info_pb['c17'] = self.avg(pb_his_high)
                stock_info_pb['d17'] = self.avg(pb_his_low)
            else:
                number = int(len(html_pb_history) / 4)
                for i in range(0, int(len(html_pb_history) / 4)):
                    get_pb_history(i)
                row_avg = 'b' + str(17 - 10 + number)
                row_high = 'c' + str(17 - 10 + number)
                row_low = 'd' + str(17 - 10 + number)
                stock_info_pb[row_avg] = self.avg(pb_his_avg)
                stock_info_pb[row_high] = self.avg(pb_his_high)
                stock_info_pb[row_low] = self.avg(pb_his_low)
        return stock_info_pb

    @staticmethod
    def avg(data):
        summary = 0
        for i in data:
            summary = summary + float(i)
        avg_data = summary / len(data)
        return round(avg_data, 2)

    def get_free_cash_flow(self, code):
        fcf_dict = {}
        fcf_url = 'https://api.wayougou.com/api/ratios/stock?code=%s&name=自由现金流&chart=自由现金流&span=10' % code
        # 'https://api.wayougou.com/api/ratios/stock?code=600006&name=自由现金流&chart=自由现金流span=10'
        data = self.urlparser.soup_request(fcf_url)
        data = json.loads(str(data))
        dates = data['自由现金流']['dates']
        dates.reverse()
        values = data['自由现金流']['values']
        values.reverse()
        for i in range(len(dates)):
            row_date = chr(97) + str(3 + i)
            row_value = chr(98) + str(3 + i)
            fcf_dict[row_date] = dates[i]
            fcf_dict[row_value] = values[i]
        return fcf_dict

    def excel(self, data1, data2, data3, code, number):
        sheet_dict = {'pe': data1, 'pb': data2, 'cash_flow': data3}
        wb = load_workbook(self.template)
        for sheet_key in sheet_dict.keys():
            data = sheet_dict[sheet_key]
            if len(data) != 0:
                ws = wb[sheet_key]
                for value_key in data.keys():
                    ws[value_key] = str(data[value_key])
                if sheet_key != 'cash_flow':
                    bd = Side(style='thin', color='000000')
                    for j in range(97, 107):
                        for i in range(1, number + 1):
                            col = chr(j) + str(i)
                            c = ws[col]
                            c.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        wb.save('%s/%s.xlsx' % (self.pe_folder, code))

    def main(self, code):
        print(code)
        data_pe_info, data_pb_info, data_fcf_info, line_number = self.get_pe_info(code)
        self.excel(data_pe_info, data_pb_info, data_fcf_info, code, line_number)


if __name__ == '__main__':
    app = ENiu()
    # app.main('000001')
    p = Pool()
    for stock_code in app.iterator_stock_code():
        p.apply_async(app.main, (stock_code,))
        # app.main(stock_code)
    p.close()
    p.join()
