from value_investment import ValuePrice
from multiprocessing import Pool
from openpyxl.styles import Border, Side, PatternFill
from openpyxl import load_workbook
from datetime import datetime


class Stock(object):
    def __init__(self):
        self.pe_path = 'data/pe'
        self.template = 'template/template_value_investment.xlsx'
        self.value_price = ValuePrice()
        self.stock_info = []

    @staticmethod
    def get_stock_codes():
        codes = []
        with open('stock.txt', 'r') as f:
            lines = f.readlines()
        for line in lines:
            codes.append(str(line).strip())
        return codes

    def iterator_stock_code(self):
        codes = self.get_stock_codes()
        for code in codes:
            yield code

    def callback(self, stock_info):
        if stock_info is not None:
            self.stock_info.append(stock_info)
        return

    def get_stock_info(self, code):
        # print(code)
        stock_info = self.value_price.estimate_value_price(code)
        # if stock_info is not None:
        #     self.stock_info.append(stock_info)
        return stock_info

    def excel_write(self):
        wb = load_workbook(self.template)
        ws = wb['A']
        for i in range(len(self.stock_info)):
            for j in range(len(self.stock_info[i])):
                row = chr(97 + j) + str(i + 2)
                ws[row] = self.stock_info[i][j]
        bd = Side(style='thin', color='000000')
        for m in range(97, 97 + len(self.stock_info[0])):
            for n in range(0, len(self.stock_info) + 1):
                row = chr(m) + str(n + 1)
                p = ws[row]
                p.border = Border(top=bd, left=bd, right=bd, bottom=bd)
        fill = PatternFill("solid", fgColor="7CCD7C")
        row_num = ws.max_row
        for i in range(2, row_num):
            pe_row = 'i' + str(i)
            pe = ws[pe_row].value
            low_pe_row = 'j' + str(i)
            low_pe = ws[low_pe_row].value
            if pe < low_pe:
                name_row = 'b' + str(i)
                row = ws[name_row]
                row.fill = fill
        wb.save('stock_%s.xlsx' % datetime.now().strftime('%Y%m%d'))

    def main(self):
        p = Pool()
        for code in self.iterator_stock_code():
            # self.get_stock_info(code)
            p.apply_async(self.get_stock_info, (code,), callback=self.callback)
        p.close()
        p.join()
        self.excel_write()


if __name__ == '__main__':
    app = Stock()
    app.main()
