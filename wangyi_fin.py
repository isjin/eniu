from fun.url_parser import UrlParser
import re
from openpyxl import load_workbook
import os
from multiprocessing import Pool


class Financial(object):
    def __init__(self):
        self.template = None
        self.parser = UrlParser()
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/70.0.3538.77 Safari/537.36',
        }
        self.template = 'template/template_wangyi_fin.xlsx'
        self.store_path = 'data/report'

    @staticmethod
    def get_stock_codes():
        codes = []
        with open('stock.txt', 'r') as f:
            lines = f.readlines()
        for line in lines:
            codes.append(str(line).strip())
        return codes

    def iterator_stock_code(self):
        codes = self.get_stock_codes()[:]
        for code in codes:
            yield code

    def report(self, url, code):
        def split_data(data, element, row_line):
            soup = self.parser.lxml_html(data)
            elements = soup.find_all(element)
            if 'class' in str(elements[0]):
                elements = elements[1:]
            if len(elements) > 10:
                for i in range(10):
                    row = chr(98 + i) + str(row_line)
                    dict_report[row] = elements[i].string
            else:
                for i in range(len(elements)):
                    row = chr(98 + i) + str(row_line)
                    dict_report[row] = elements[i].string

        def splite_table(classname):
            html_tables = html_fin.find_all('table', class_=classname)
            for html_table in html_tables:
                html_table = self.parser.lxml_html(html_table)
                html_table_trs = html_table.find_all('tr')
                if len(dict_report) == 1:
                    split_data(html_table_trs[0], 'th', 2)
                    for tr in (html_table_trs[1:]):
                        last_key = list(dict_report.keys())[-1]
                        last_line_num = re.search(r'\d+', str(last_key)).group()
                        split_data(tr, 'td', int(last_line_num) + 1)
                else:
                    for tr in (html_table_trs[1:]):
                        last_key = list(dict_report.keys())[-1]
                        last_line_num = re.search(r'\d+', str(last_key)).group()
                        split_data(tr, 'td', int(last_line_num) + 1)

        dict_report = {}

        html_fin = self.parser.soup_request(url, headers=self.headers)
        name = html_fin.find_all('h1', class_='name')[0].a.string
        name = str(name).replace(' ', '')
        dict_report['b1'] = '%s_%s' % (code, name)
        splite_table('table_bg001 border_box limit_sale scr_table')
        splite_table('table_bg001 border_box fund_analys')
        return dict_report

    def shareholder(self, url):
        dict_shareholder_table = {}
        dict_shareholder_ten = {}
        html = self.parser.soup_request(url)
        html_share_stock = html.find_all('table', class_='table_bg001 border_box')
        html_share_stock = self.parser.lxml_html(html_share_stock)
        html_share_stock = html_share_stock.find_all('td')
        for x in range(int(len(html_share_stock) / 3)):
            for y in range(3):
                row_share_stock = chr(97 + y) + str(2 + x)
                dict_shareholder_table[row_share_stock] = html_share_stock[x * 3 + y].string
        html_sh_table = html.find('table', class_='table_bg001 border_box gudong_table')
        html_sh_count = self.parser.lxml_html(html_sh_table)
        html_sh_count = html_sh_count.find_all('td')
        for i in range(int(len(html_sh_count) / 5)):
            for j in range(5):
                row_shareholder_table = chr(97 + j) + str(12 + i)
                dict_shareholder_table[row_shareholder_table] = html_sh_count[i * 5 + j].string

        html_sh_ten = html.find_all('table', class_='table_bg001 border_box limit_sale')[1:]
        for m in range(2):
            html_shareholder_ten = self.parser.lxml_html(html_sh_ten)
            html_shareholder_ten = html_shareholder_ten.find_all('td')
            for o in range(int(len(html_shareholder_ten) / 8)):
                for q in range(4):
                    if m == 0:
                        row_shareholder_ten = chr(97 + q) + str(2 + o)
                    else:
                        row_shareholder_ten = chr(97 + q) + str(16 + o)
                    dict_shareholder_ten[row_shareholder_ten] = html_shareholder_ten[o * 4 + q].string
        return dict_shareholder_table, dict_shareholder_ten

    # def stock_cash_flow(self, data, code):
    #     sohu_url = 'http://q.stock.sohu.com/cn/%s/cwzb.shtml' % code
    #     html = self.parser.soup_request(sohu_url)
    #     html = html.find_all('tr', class_='e4')[2]
    #     html = self.parser.lxml_html(html)
    #     html = html.find_all('td')[0]
    #     cash_flow = str(html.string).strip()
    #     data['b5'] = cash_flow
    #     return data

    def main(self, code):
        def excel_write(ws, dict_data):
            for value_key in dict_data.keys():
                ws[value_key] = str(dict_data[value_key])

        print(code)
        url_report_quarter = 'http://quotes.money.163.com/f10/zycwzb_%s,report.html' % code
        url_report_year = 'http://quotes.money.163.com/f10/zycwzb_%s,year.html' % code
        url_shareholder = 'http://quotes.money.163.com/f10/gdfx_%s.html#01d02' % code
        dict_quarter = self.report(url_report_quarter, code)
        # dict_quarter = self.stock_cash_flow(dict_quarter, code)
        dict_year = self.report(url_report_year, code)
        dict_shareholder_table, dict_shareholder_ten = self.shareholder(url_shareholder)
        wb = load_workbook(self.template)
        ws1 = wb['report_quarter']
        ws2 = wb['report_year']
        ws3 = wb['shareholder_table']
        ws4 = wb['shareholder_ten']
        excel_write(ws1, dict_quarter)
        excel_write(ws2, dict_year)
        excel_write(ws3, dict_shareholder_table)
        excel_write(ws4, dict_shareholder_ten)
        file_name = '%s.xlsx' % code
        file_path = os.path.join(self.store_path, file_name)
        wb.save(file_path)


if __name__ == '__main__':
    app = Financial()
    p = Pool()
    for stock_code in app.iterator_stock_code():
    # for stock_code in ['000002']:
    #     app.main(stock_code)
        p.apply_async(app.main, (stock_code,))
    p.close()
    p.join()
    # url = 'http://quotes.money.163.com/f10/gdfx_600006.html#01d02'
    # app.shareholder(url)
